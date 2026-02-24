"""Leave management API routes"""
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, status, Query, Request
from fastapi.responses import StreamingResponse, JSONResponse
from sqlalchemy import select, func, and_
from sqlalchemy.ext.asyncio import AsyncSession
import csv
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.core.database import get_db
from app.core.permissions import require_permission
from app.core.query_utils import (
    apply_search, apply_ordering, apply_expansion, parse_expand_param
)
from app.audit_app.services import AuditService
from app.user_app.models import User
from app.conge_app.models import TypeConge, DemandeConge, SoldeConge, HistoriqueConge
from app.conge_app.schemas import (
    TypeCongeCreate, TypeCongeUpdate, TypeCongeResponse, PaginatedTypeConge,
    DemandeCongeCreate, DemandeCongeUpdate, DemandeCongeResponse,
    PaginatedDemandeConge, ApproveRejectRequest,
    SoldeCongeCreate, SoldeCongeUpdate, SoldeCongeResponse,
    PaginatedSoldeConge, BulkCreateSoldeRequest,
    HistoriqueCongeResponse, PaginatedHistoriqueConge,
    CongeStatsResponse, EmployeStatsResponse, ServiceStatsResponse
)
from app.conge_app.services.demande_service import DemandeCongeService
from app.conge_app.services.validation_service import ValidationService
from app.conge_app.services.solde_service import SoldeCongeService

router = APIRouter(prefix="/api/conge", tags=["Congé Management"])


# ============================================================================
# TypeConge Routes
# ============================================================================

@router.get("/types", response_model=PaginatedTypeConge)
async def list_type_conge(
    db: AsyncSession = Depends(get_db),
    _current_user: User = Depends(require_permission("conge", "view")),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    no_pagination: bool = Query(False),
    search: Optional[str] = Query(None),
    ordering: Optional[str] = Query(None),
    expand: Optional[str] = Query(None)
):
    """
    List all leave types with pagination, search, and expand support.

    Retrieve a paginated list of all leave types (TypeConge) in the system.
    Supports searching, ordering, and expanding related entities.

    **Query Parameters:**
    - **skip**: Number of records to skip (default: 0)
    - **limit**: Maximum number of records to return (default: 100, max: 500)
    - **no_pagination**: If true, returns all records without pagination
    - **search**: Search term to filter by name, code, or description
    - **ordering**: Field to order by (prefix with '-' for descending, e.g., '-nom')
    - **expand**: Comma-separated list of relations to expand

    **Example Response:**
    ```json
    {
        "items": [
            {
                "id": 1,
                "nom": "Congé Annuel",
                "code": "CA",
                "description": "Congé annuel payé",
                "nb_jours_max": 30,
                "est_paye": true,
                "necessite_justificatif": false,
                "couleur": "#4CAF50",
                "actif": true
            }
        ],
        "total": 1,
        "skip": 0,
        "limit": 100
    }
    ```

    **Error Responses:**
    - **401**: Unauthorized - Invalid or missing authentication
    - **403**: Forbidden - Insufficient permissions
    """
    query = select(TypeConge)
    search_fields = ['nom', 'code', 'description']

    if search:
        query = apply_search(query, TypeConge, search_fields, search)

    if ordering:
        query = apply_ordering(query, TypeConge, ordering)
    else:
        query = query.order_by(TypeConge.nom.asc())

    expand_fields = parse_expand_param(expand)
    if expand_fields:
        query = apply_expansion(query, TypeConge, expand_fields)

    count_query = select(func.count()).select_from(TypeConge)
    if search:
        count_query = apply_search(count_query, TypeConge, search_fields, search)

    result = await db.execute(count_query)
    total = result.scalar()

    if not no_pagination:
        query = query.offset(skip).limit(limit)
    else:
        skip = 0
        limit = total

    result = await db.execute(query)
    types = result.scalars().all()

    return PaginatedTypeConge(
        items=[TypeCongeResponse.model_validate(t) for t in types],
        total=total,
        skip=skip,
        limit=limit
    )


@router.post("/types", response_model=TypeCongeResponse)
async def create_type_conge(
    type_data: TypeCongeCreate,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission("conge", "manage_types"))
):
    """
    Create a new leave type.

    Create a new leave type (TypeConge) in the system. Each leave type must have
    a unique code and defines the characteristics of a leave category.

    **Request Body:**
    ```json
    {
        "nom": "Congé Annuel",
        "code": "CA",
        "description": "Congé annuel payé réglementaire",
        "nb_jours_max": 30,
        "est_paye": true,
        "necessite_justificatif": false,
        "couleur": "#4CAF50",
        "actif": true
    }
    ```

    **Example Response:**
    ```json
    {
        "id": 1,
        "nom": "Congé Annuel",
        "code": "CA",
        "description": "Congé annuel payé réglementaire",
        "nb_jours_max": 30,
        "est_paye": true,
        "necessite_justificatif": false,
        "couleur": "#4CAF50",
        "actif": true,
        "created_at": "2024-01-15T10:30:00",
        "updated_at": "2024-01-15T10:30:00"
    }
    ```

    **Error Responses:**
    - **400**: Bad Request - Leave type with this code already exists
    - **401**: Unauthorized - Invalid or missing authentication
    - **403**: Forbidden - Insufficient permissions (requires 'manage_types')
    - **422**: Validation Error - Invalid request body
    """
    existing_query = select(TypeConge).where(TypeConge.code == type_data.code)
    result = await db.execute(existing_query)
    existing = result.scalar_one_or_none()

    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Type de congé avec le code '{type_data.code}' existe déjà"
        )

    new_type = TypeConge(**type_data.model_dump())
    db.add(new_type)
    await db.commit()
    await db.refresh(new_type)

    await AuditService.log_model_change(
        db=db,
        user=current_user,
        instance=new_type,
        action="CREATE",
        request=request
    )

    return TypeCongeResponse.model_validate(new_type)


@router.get("/types/{type_id}", response_model=TypeCongeResponse)
async def get_type_conge(
    type_id: int,
    db: AsyncSession = Depends(get_db),
    _current_user: User = Depends(require_permission("conge", "view")),
    expand: Optional[str] = Query(None)
):
    """Get a specific TypeConge by ID."""
    query = select(TypeConge).where(TypeConge.id == type_id)

    expand_fields = parse_expand_param(expand)
    if expand_fields:
        query = apply_expansion(query, TypeConge, expand_fields)

    result = await db.execute(query)
    type_conge = result.scalar_one_or_none()

    if not type_conge:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Type de congé non trouvé"
        )

    return TypeCongeResponse.model_validate(type_conge)


@router.put("/types/{type_id}", response_model=TypeCongeResponse)
async def update_type_conge(
    type_id: int,
    type_data: TypeCongeUpdate,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission("conge", "manage_types"))
):
    """Update a TypeConge."""
    query = select(TypeConge).where(TypeConge.id == type_id)
    result = await db.execute(query)
    type_conge = result.scalar_one_or_none()

    if not type_conge:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Type de congé non trouvé"
        )

    old_values = AuditService._extract_model_values(type_conge)
    update_data = type_data.model_dump(exclude_unset=True)

    for field, value in update_data.items():
        setattr(type_conge, field, value)

    await db.commit()
    await db.refresh(type_conge)

    await AuditService.log_model_change(
        db=db,
        user=current_user,
        instance=type_conge,
        action="UPDATE",
        old_values=old_values,
        request=request
    )

    return TypeCongeResponse.model_validate(type_conge)


@router.delete("/types/{type_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_type_conge(
    type_id: int,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission("conge", "manage_types"))
):
    """Delete a TypeConge."""
    query = select(TypeConge).where(TypeConge.id == type_id)
    result = await db.execute(query)
    type_conge = result.scalar_one_or_none()

    if not type_conge:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Type de congé non trouvé"
        )

    active_demandes_query = select(func.count()).select_from(
        DemandeConge
    ).where(
        and_(
            DemandeConge.type_conge_id == type_id,
            DemandeConge.statut.notin_(['CANCELLED', 'REJECTED'])
        )
    )

    result = await db.execute(active_demandes_query)
    active_count = result.scalar()

    if active_count > 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"Impossible de supprimer ce type de congé. "
                f"{active_count} demande(s) active(s) y font référence."
            )
        )

    old_values = AuditService._extract_model_values(type_conge)
    await db.delete(type_conge)
    await db.commit()

    await AuditService.log_action(
        db=db,
        user=current_user,
        action="DELETE",
        resource_type="cg_type_conge",
        resource_id=str(type_id),
        old_values=old_values,
        request=request
    )

    return None


# ============================================================================
# DemandeConge Routes
# ============================================================================

@router.get("/demandes/export")
async def export_demandes(
    export_format: str = Query("json", pattern="^(json|csv|excel)$"),
    request: Request = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission("conge", "export")),
    employe_id: Optional[int] = Query(None),
    type_conge_id: Optional[int] = Query(None),
    statut: Optional[str] = Query(None),
    date_debut: Optional[str] = Query(None),
    date_fin: Optional[str] = Query(None),
    expand: Optional[str] = Query(None)
):
    """
    Export leave requests data in JSON, CSV, or Excel format.

    Export all leave requests matching the specified filters in the requested format.
    Supports JSON (with metadata), CSV (UTF-8 with BOM), and Excel (XLSX) formats.

    **Query Parameters:**
    - **export_format**: Export format - 'json', 'csv', or 'excel' (default: 'json')
    - **employe_id**: Filter by employee ID
    - **type_conge_id**: Filter by leave type ID
    - **statut**: Filter by status (PENDING, IN_PROGRESS, APPROVED, REJECTED, CANCELLED)
    - **date_debut**: Filter by start date (format: YYYY-MM-DD)
    - **date_fin**: Filter by end date (format: YYYY-MM-DD)
    - **expand**: Comma-separated list of relations to expand (e.g., 'employe,type_conge')

    **Example JSON Response:**
    ```json
    {
        "data": [
            {
                "id": 1,
                "employe_id": 5,
                "type_conge_id": 1,
                "date_debut": "2024-02-01",
                "date_fin": "2024-02-05",
                "est_demi_journee": false,
                "periode_demi_journee": "",
                "nb_jours_demandes": 5.0,
                "nb_jours_ouvrables": 5.0,
                "raison": "Vacances familiales",
                "statut": "APPROVED",
                "niveau_validation_actuel": 2,
                "date_soumission": "2024-01-15T10:30:00",
                "date_decision_finale": "2024-01-16T14:20:00",
                "employe_nom": "Dupont Jean",
                "type_conge_nom": "Congé Annuel"
            }
        ],
        "total": 1,
        "exported_at": "2024-01-20T15:30:00",
        "filters": {
            "employe_id": 5,
            "statut": "APPROVED"
        }
    }
    ```

    **CSV Format:**
    - Returns a downloadable CSV file with UTF-8 BOM encoding
    - Filename: demandes_conge_YYYYMMDD_HHMMSS.csv
    - Includes all fields and expanded relations

    **Excel Format:**
    - Returns a downloadable XLSX file with formatted headers
    - Filename: demandes_conge_YYYYMMDD_HHMMSS.xlsx
    - Styled header row with blue background
    - Auto-adjusted column widths

    **Error Responses:**
    - **401**: Unauthorized - Invalid or missing authentication
    - **403**: Forbidden - Insufficient permissions (requires 'export')
    - **422**: Validation Error - Invalid export format or query parameters
    """
    filters = {}
    if employe_id:
        filters['employe_id'] = employe_id
    if type_conge_id:
        filters['type_conge_id'] = type_conge_id
    if statut:
        filters['statut'] = statut
    if date_debut:
        filters['date_debut'] = date_debut
    if date_fin:
        filters['date_fin'] = date_fin

    # Get all demandes with expand support
    demandes, total = await DemandeCongeService.list_demandes(
        filters=filters,
        expand=expand,
        skip=0,
        limit=None,
        no_pagination=True,
        search=None,
        ordering=None,
        db=db
    )

    # Log export action
    await AuditService.log_action(
        db=db,
        user=current_user,
        action="EXPORT",
        resource_type="cg_demande_conge",
        resource_id=None,
        new_values={
            "format": export_format,
            "count": total,
            "filters": filters
        },
        request=request
    )

    # Prepare data for export
    export_data = []
    for demande in demandes:
        row = {
            "id": demande.id,
            "employe_id": demande.employe_id,
            "type_conge_id": demande.type_conge_id,
            "date_debut": str(demande.date_debut),
            "date_fin": str(demande.date_fin),
            "est_demi_journee": demande.est_demi_journee,
            "periode_demi_journee": demande.periode_demi_journee or "",
            "nb_jours_demandes": float(demande.nb_jours_demandes),
            "nb_jours_ouvrables": float(demande.nb_jours_ouvrables),
            "raison": demande.raison,
            "statut": demande.statut,
            "niveau_validation_actuel": demande.niveau_validation_actuel,
            "date_soumission": str(demande.date_soumission),
            "date_decision_finale": (
                str(demande.date_decision_finale)
                if demande.date_decision_finale else ""
            ),
        }

        # Add expanded relations if available
        if hasattr(demande, 'employe') and demande.employe:
            row["employe_nom"] = f"{demande.employe.nom} {demande.employe.prenom}"
        if hasattr(demande, 'type_conge') and demande.type_conge:
            row["type_conge_nom"] = demande.type_conge.nom

        export_data.append(row)

    # Generate export based on format
    if export_format == "json":
        return JSONResponse(
            content={
                "data": export_data,
                "total": total,
                "exported_at": datetime.now().isoformat(),
                "filters": filters
            }
        )

    elif export_format == "csv":
        # Create CSV in memory
        output = io.StringIO()
        if export_data:
            writer = csv.DictWriter(output, fieldnames=export_data[0].keys())
            writer.writeheader()
            writer.writerows(export_data)

        csv_content = output.getvalue()
        output.close()

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"demandes_conge_{timestamp}.csv"

        return StreamingResponse(
            io.BytesIO(csv_content.encode('utf-8-sig')),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    elif export_format == "excel":
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Demandes de Congé"

        # Define headers
        headers = [
            "ID", "Employé ID", "Type Congé ID", "Date Début", "Date Fin",
            "Demi-Journée", "Période", "Jours Demandés", "Jours Ouvrables",
            "Raison", "Statut", "Niveau Validation", "Date Soumission",
            "Date Décision"
        ]

        # Add expanded relation headers if available
        if export_data and "employe_nom" in export_data[0]:
            headers.append("Employé")
        if export_data and "type_conge_nom" in export_data[0]:
            headers.append("Type de Congé")

        # Style header row
        header_fill = PatternFill(
            start_color="366092",
            end_color="366092",
            fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add data rows
        for row_num, demande_data in enumerate(export_data, 2):
            ws.cell(row=row_num, column=1, value=demande_data["id"])
            ws.cell(row=row_num, column=2, value=demande_data["employe_id"])
            ws.cell(row=row_num, column=3, value=demande_data["type_conge_id"])
            ws.cell(row=row_num, column=4, value=demande_data["date_debut"])
            ws.cell(row=row_num, column=5, value=demande_data["date_fin"])
            ws.cell(
                row=row_num, column=6,
                value="Oui" if demande_data["est_demi_journee"] else "Non"
            )
            ws.cell(
                row=row_num, column=7,
                value=demande_data["periode_demi_journee"]
            )
            ws.cell(
                row=row_num, column=8,
                value=demande_data["nb_jours_demandes"]
            )
            ws.cell(
                row=row_num, column=9,
                value=demande_data["nb_jours_ouvrables"]
            )
            ws.cell(row=row_num, column=10, value=demande_data["raison"])
            ws.cell(row=row_num, column=11, value=demande_data["statut"])
            ws.cell(
                row=row_num, column=12,
                value=demande_data["niveau_validation_actuel"]
            )
            ws.cell(
                row=row_num, column=13,
                value=demande_data["date_soumission"]
            )
            ws.cell(
                row=row_num, column=14,
                value=demande_data["date_decision_finale"]
            )

            # Add expanded relation data if available
            col_offset = 14
            if "employe_nom" in demande_data:
                ws.cell(
                    row=row_num, column=col_offset + 1,
                    value=demande_data.get("employe_nom", "")
                )
                col_offset += 1
            if "type_conge_nom" in demande_data:
                ws.cell(
                    row=row_num, column=col_offset + 1,
                    value=demande_data.get("type_conge_nom", "")
                )

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        excel_output = io.BytesIO()
        wb.save(excel_output)
        excel_output.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"demandes_conge_{timestamp}.xlsx"

        return StreamingResponse(
            excel_output,
            media_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )


@router.get("/demandes", response_model=PaginatedDemandeConge)
async def list_demande_conge(
    db: AsyncSession = Depends(get_db),
    _current_user: User = Depends(require_permission("conge", "view")),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    no_pagination: bool = Query(False),
    search: Optional[str] = Query(None),
    ordering: Optional[str] = Query(None),
    expand: Optional[str] = Query(None),
    employe_id: Optional[int] = Query(None),
    type_conge_id: Optional[int] = Query(None),
    statut: Optional[str] = Query(None),
    date_debut: Optional[str] = Query(None),
    date_fin: Optional[str] = Query(None)
):
    """
    List all leave requests with pagination, search, filters, and expand support.

    Retrieve a paginated list of leave requests (DemandeConge) with advanced filtering
    capabilities. Supports filtering by employee, leave type, status, and date range.

    **Query Parameters:**
    - **skip**: Number of records to skip (default: 0)
    - **limit**: Maximum number of records to return (default: 100, max: 500)
    - **no_pagination**: If true, returns all records without pagination
    - **search**: Search term to filter by reason
    - **ordering**: Field to order by (prefix with '-' for descending)
    - **expand**: Comma-separated list of relations to expand
    - **employe_id**: Filter by employee ID
    - **type_conge_id**: Filter by leave type ID
    - **statut**: Filter by status
    - **date_debut**: Filter by start date (format: YYYY-MM-DD)
    - **date_fin**: Filter by end date (format: YYYY-MM-DD)

    **Example Response:**
    ```json
    {
        "items": [
            {
                "id": 1,
                "employe_id": 5,
                "type_conge_id": 1,
                "date_debut": "2024-02-01",
                "date_fin": "2024-02-05",
                "est_demi_journee": false,
                "periode_demi_journee": null,
                "nb_jours_demandes": 5.0,
                "nb_jours_ouvrables": 5.0,
                "raison": "Vacances familiales",
                "statut": "APPROVED",
                "niveau_validation_actuel": 2,
                "date_soumission": "2024-01-15T10:30:00",
                "date_decision_finale": "2024-01-16T14:20:00"
            }
        ],
        "total": 1,
        "skip": 0,
        "limit": 100
    }
    ```

    **Error Responses:**
    - **401**: Unauthorized - Invalid or missing authentication
    - **403**: Forbidden - Insufficient permissions
    - **422**: Validation Error - Invalid query parameters
    """
    filters = {}
    if employe_id:
        filters['employe_id'] = employe_id
    if type_conge_id:
        filters['type_conge_id'] = type_conge_id
    if statut:
        filters['statut'] = statut
    if date_debut:
        filters['date_debut'] = date_debut
    if date_fin:
        filters['date_fin'] = date_fin

    demandes, total = await DemandeCongeService.list_demandes(
        filters=filters,
        expand=expand,
        skip=skip if not no_pagination else 0,
        limit=limit if not no_pagination else None,
        no_pagination=no_pagination,
        search=search,
        ordering=ordering,
        db=db
    )

    return PaginatedDemandeConge(
        items=[DemandeCongeResponse.model_validate(d) for d in demandes],
        total=total,
        skip=skip if not no_pagination else 0,
        limit=limit if not no_pagination else total
    )


@router.post("/demandes", response_model=DemandeCongeResponse)
async def create_demande_conge(
    demande_data: DemandeCongeCreate,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission("conge", "create"))
):
    """
    Create a new leave request.

    Submit a new leave request (DemandeConge) for an employee. The system will
    automatically calculate working days, validate against available balance,
    check for overlapping requests, and initialize the validation workflow.

    **Request Body:**
    ```json
    {
        "employe_id": 5,
        "type_conge_id": 1,
        "date_debut": "2024-02-01",
        "date_fin": "2024-02-05",
        "est_demi_journee": false,
        "periode_demi_journee": null,
        "raison": "Vacances familiales"
    }
    ```

    **Example for Half-Day Leave:**
    ```json
    {
        "employe_id": 5,
        "type_conge_id": 1,
        "date_debut": "2024-02-01",
        "date_fin": "2024-02-01",
        "est_demi_journee": true,
        "periode_demi_journee": "MATIN",
        "raison": "Rendez-vous médical"
    }
    ```

    **Example Response:**
    ```json
    {
        "id": 1,
        "employe_id": 5,
        "type_conge_id": 1,
        "date_debut": "2024-02-01",
        "date_fin": "2024-02-05",
        "est_demi_journee": false,
        "periode_demi_journee": null,
        "nb_jours_demandes": 5.0,
        "nb_jours_ouvrables": 5.0,
        "raison": "Vacances familiales",
        "statut": "PENDING",
        "niveau_validation_actuel": 1,
        "date_soumission": "2024-01-15T10:30:
    demande = await DemandeCongeService.create_demande(
        demande_data=demande_data,
        db=db
    )

    await AuditService.log_model_change(
        db=db,
        user=current_user,
        instance=demande,
        action="CREATE",
        request=request
    )

    return DemandeCongeResponse.model_validate(demande)
ent balance, overlapping dates, etc.)
    - **401**: Unauthorized - Invalid or missing authentication
    - **403**: Forbidden - Insufficient permissions
    - **404**: Not Found - Employee or leave type not found
    - **422**: Validation Error - Invalid request body
    """

@router.get("/demandes/{demande_id}", response_model=DemandeCongeResponse)
async def get_demande_conge(
    demande_id: int,
    db: AsyncSession = Depends(get_db),
    _current_user: User = Depends(require_permission("conge", "view")),
    expand: Optional[str] = Query(None)
):
    """Get a specific DemandeConge by ID."""
    query = select(DemandeConge).where(DemandeConge.id == demande_id)

    expand_fields = parse_expand_param(expand)
    if expand_fields:
        query = apply_expansion(query, DemandeConge, expand_fields)

    result = await db.execute(query)
    demande = result.scalar_one_or_none()

    if not demande:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Demande de congé non trouvée"
        )

    return DemandeCongeResponse.model_validate(demande)


@router.put("/demandes/{demande_id}", response_model=DemandeCongeResponse)
async def update_demande_conge(
    demande_id: int,
    demande_data: DemandeCongeUpdate,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission("conge", "update"))
):
    """Update a DemandeConge (only if status is PENDING)."""
    demande = await DemandeCongeService.update_demande(
        demande_id=demande_id,
        demande_data=demande_data,
        db=db
    )

    await AuditService.log_model_change(
        db=db,
        user=current_user,
        instance=demande,
        action="UPDATE",
        request=request
    )

    return DemandeCongeResponse.model_validate(demande)


@router.delete("/demandes/{demande_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_demande_conge(
    demande_id: int,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission("conge", "delete"))
):
    """Cancel a DemandeConge."""
    demande = await DemandeCongeService.cancel_demande(
        demande_id=demande_id,
        user_id=current_user.id,
        db=db
    )

    await AuditService.log_action(
        db=db,
        user=current_user,
        action="DELETE",
        resource_type="cg_demande_conge",
        resource_id=str(demande_id),
        old_values=AuditService._extract_model_values(demande),
        request=request
    )

    return None


@router.post("/demandes/{demande_id}/approve", response_model=DemandeCongeResponse)
async def approve_demande_conge(
    demande_id: int,
    approve_data: ApproveRejectRequest,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission("conge", "approve"))
):
    """
    Approve a leave request at the current validation level.

    Approve a leave request (DemandeConge) at the current validation level.
    If this is the final validation level, the request status will be set to APPROVED
    and the employee's leave balance will be updated. Otherwise, the request advances
    to the next validation level.

    **Path Parameters:**
    - **demande_id**: ID of the leave request to approve

    **Request Body:**
    ```json
    {
        "commentaire": "Approuvé - période de faible activité"
    }
    ```

    **Example Response:**
    ```json
    {
        "id": 1,
        "employe_id": 5,
        "type_conge_id": 1,
        "date_debut": "2024-02-01",
        "date_fin": "2024-02-05",
        "est_demi_journee": false,
        "periode_demi_journee": null,
        "nb_jours_demandes": 5.0,
        "nb_jours_ouvrables": 5.0,
        "raison": "Vacances familiales",
        "statut": "APPROVED",
        "niveau_validation_actuel": 2,
        "date_soumission": "2024-01-15T10:30:00",
        "date_decision_finale": "2024-01-16T14:20:00",
        "created_at": "2024-01-15T10:30:00",
        "updated_at": "2024-01-16T14:20:00"
    }
    ```

    **Workflow:**
    1. Validates that the request is in PENDING or IN_PROGRESS status
    2. Verifies the current user has approval rights at this validation level
    3. Records the approval in the validation history
    4. If final level: Sets status to APPROVED and updates leave balance
    5. If not final: Advances to next validation level (status: IN_PROGRESS)

    **Error Responses:**
    - **400**: Bad Request - Request not in valid status or user not authorized for this level
    - **401**: Unauthorized - Invalid or missing authentication
    - **403**: Forbidden - Insufficient permissions (requires 'approve')
    - **404**: Not Found - Leave request not found
    - **422**: Validation Error - Invalid request body
    """
    demande = await ValidationService.approve_at_level(
        demande_id=demande_id,
        valideur_id=current_user.id,
        commentaire=approve_data.commentaire,
        db=db
    )

    await AuditService.log_action(
        db=db,
        user=current_user,
        action="APPROVE",
        resource_type="cg_demande_conge",
        resource_id=str(demande_id),
        new_values={
            "statut": demande.statut,
            "niveau_validation_actuel": demande.niveau_validation_actuel,
            "commentaire": approve_data.commentaire
        },
        request=request
    )

    return DemandeCongeResponse.model_validate(demande)


@router.post("/demandes/{demande_id}/reject", response_model=DemandeCongeResponse)
async def reject_demande_conge(
    demande_id: int,
    reject_data: ApproveRejectRequest,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission("conge", "approve"))
):
    """
    Reject a leave request at the current validation level.

    Reject a leave request (DemandeConge) at the current validation level.
    The request status will be set to REJECTED and no further validation is possible.
    The leave balance is not affected by rejection.

    **Path Parameters:**
    - **demande_id**: ID of the leave request to reject

    **Request Body:**
    ```json
    {
        "commentaire": "Rejeté - période de forte activité, report nécessaire"
    }
    ```

    **Example Response:**
    ```json
    {
        "id": 1,
        "employe_id": 5,
        "type_conge_id": 1,
        "date_debut": "2024-02-01",
        "date_fin": "2024-02-05",
        "est_demi_journee": false,
        "periode_demi_journee": null,
        "nb_jours_demandes": 5.0,
        "nb_jours_ouvrables": 5.0,
        "raison": "Vacances familiales",
        "statut": "REJECTED",
        "niveau_validation_actuel": 1,
        "date_soumission": "2024-01-15T10:30:00",
        "date_decision_finale": "2024-01-16T09:15:00",
        "created_at": "2024-01-15T10:30:00",
        "updated_at": "2024-01-16T09:15:00"
    }
    ```

    **Workflow:**
    1. Validates that the request is in PENDING or IN_PROGRESS status
    2. Verifies the current user has approval rights at this validation level
    3. Records the rejection in the validation history with comment
    4. Sets status to REJECTED and records decision date
    5. No leave balance is deducted

    **Error Responses:**
    - **400**: Bad Request - Request not in valid status or user not authorized for this level
    - **401**: Unauthorized - Invalid or missing authentication
    - **403**: Forbidden - Insufficient permissions (requires 'approve')
    - **404**: Not Found - Leave request not found
    - **422**: Validation Error - Invalid request body
    """
    demande = await ValidationService.reject_at_level(
        demande_id=demande_id,
        valideur_id=current_user.id,
        commentaire=reject_data.commentaire,
        db=db
    )

    await AuditService.log_action(
        db=db,
        user=current_user,
        action="REJECT",
        resource_type="cg_demande_conge",
        resource_id=str(demande_id),
        new_values={
            "statut": demande.statut,
            "commentaire": reject_data.commentaire
        },
        request=request
    )

    return DemandeCongeResponse.model_validate(demande)


@router.post("/demandes/{demande_id}/delegate")
async def delegate_demande_conge(
    demande_id: int,
    delegate_data: ApproveRejectRequest,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission("conge", "approve"))
):
    """Delegate validation of a DemandeConge to another user."""
    if not delegate_data.delegue_a_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="delegue_a_id est requis pour la délégation"
        )

    historique = await ValidationService.delegate_validation(
        demande_id=demande_id,
        valideur_id=current_user.id,
        delegue_a_id=delegate_data.delegue_a_id,
        commentaire=delegate_data.commentaire,
        db=db
    )

    await AuditService.log_action(
        db=db,
        user=current_user,
        action="DELEGATE",
        resource_type="cg_demande_conge",
        resource_id=str(demande_id),
        new_values={
            "delegue_a_id": delegate_data.delegue_a_id,
            "commentaire": delegate_data.commentaire
        },
        request=request
    )

    return {
        "message": "Validation déléguée avec succès",
        "historique_id": historique.id,
        "delegue_a_id": delegate_data.delegue_a_id
    }


# ============================================================================
# SoldeConge Routes
# ============================================================================

@router.get("/soldes", response_model=PaginatedSoldeConge)
async def list_solde_conge(
    db: AsyncSession = Depends(get_db),
    _current_user: User = Depends(require_permission("conge", "view")),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    no_pagination: bool = Query(False),
    ordering: Optional[str] = Query(None),
    expand: Optional[str] = Query(None),
    employe_id: Optional[int] = Query(None),
    type_conge_id: Optional[int] = Query(None),
    annee: Optional[int] = Query(None)
):
    """
    List all leave balances with pagination, filters, and expand support.

    Retrieve a paginated list of leave balances (SoldeConge) for employees.
    Each balance tracks allocated, used, and remaining leave days for a specific
    employee, leave type, and year.

    **Query Parameters:**
    - **skip**: Number of records to skip (default: 0)
    - **limit**: Maximum number of records to return (default: 100, max: 500)
    - **no_pagination**: If true, returns all records without pagination
    - **ordering**: Field to order by (prefix with '-' for descending, e.g., '-annee')
    - **expand**: Comma-separated list of relations to expand (e.g., 'employe,type_conge')
    - **employe_id**: Filter by employee ID
    - **type_conge_id**: Filter by leave type ID
    - **annee**: Filter by year (e.g., 2024)

    **Example Response:**
    ```json
    {
        "items": [
            {
                "id": 1,
                "employe_id": 5,
                "type_conge_id": 1,
                "annee": 2024,
                "alloue": 30.0,
                "utilise": 10.0,
                "restant": 20.0,
                "created_at": "2024-01-01T00:00:00",
                "updated_at": "2024-01-15T10:30:00"
            }
        ],
        "total": 1,
        "skip": 0,
        "limit": 100
    }
    ```

    **Balance Calculation:**
    - **alloue**: Total days allocated for the year
    - **utilise**: Days used from approved leave requests
    - **restant**: Remaining days (alloue - utilise)

    **Error Responses:**
    - **401**: Unauthorized - Invalid or missing authentication
    - **403**: Forbidden - Insufficient permissions
    - **422**: Validation Error - Invalid query parameters
    """
    filters = {}
    if employe_id:
        filters['employe_id'] = employe_id
    if type_conge_id:
        filters['type_conge_id'] = type_conge_id
    if annee:
        filters['annee'] = annee

    soldes, total = await SoldeCongeService.list_soldes(
        filters=filters,
        expand=expand,
        skip=skip if not no_pagination else 0,
        limit=limit if not no_pagination else None,
        no_pagination=no_pagination,
        search=None,
        ordering=ordering,
        db=db
    )

    return PaginatedSoldeConge(
        items=[SoldeCongeResponse.model_validate(s) for s in soldes],
        total=total,
        skip=skip if not no_pagination else 0,
        limit=limit if not no_pagination else total
    )


@router.post("/soldes", response_model=SoldeCongeResponse)
async def create_solde_conge(
    solde_data: SoldeCongeCreate,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission("conge", "manage_soldes"))
):
    """
    Create a new leave balance with automatic remaining days calculation.

    Create a new leave balance (SoldeConge) for an employee. The system automatically
    calculates the remaining days based on allocated and used days. Each employee
    can have only one balance per leave type per year.

    **Request Body:**
    ```json
    {
        "employe_id": 5,
        "type_conge_id": 1,
        "annee": 2024,
        "alloue": 30.0,
        "utilise": 0.0
    }
    ```

    **Example Response:**
    ```json
    {
        "id": 1,
        "employe_id": 5,
        "type_conge_id": 1,
        "annee": 2024,
        "alloue": 30.0,
        "utilise": 0.0,
        "restant": 30.0,
        "created_at": "2024-01-01T00:00:00",
        "updated_at": "2024-01-01T00:00:00"
    }
    ```

    **Validation Rules:**
    - Employee and leave type must exist
    - Only one balance per employee/type/year combination
    - Allocated days must be >= 0
    - Used days must be >= 0 and <= allocated days
    - Remaining days calculated automatically (alloue - utilise)

    **Error Responses:**
    - **400**: Bad Request - Balance already exists for this employee/type/year
    - **401**: Unauthorized - Invalid or missing authentication
    - **403**: Forbidden - Insufficient permissions (requires 'manage_soldes')
    - **404**: Not Found - Employee or leave type not found
    - **422**: Validation Error - Invalid request body
    """
    solde = await SoldeCongeService.create_solde(
        solde_data=solde_data,
        db=db
    )

    await AuditService.log_model_change(
        db=db,
        user=current_user,
        instance=solde,
        action="CREATE",
        request=request
    )

    return SoldeCongeResponse.model_validate(solde)


@router.get("/soldes/{solde_id}", response_model=SoldeCongeResponse)
async def get_solde_conge(
    solde_id: int,
    db: AsyncSession = Depends(get_db),
    _current_user: User = Depends(require_permission("conge", "view")),
    expand: Optional[str] = Query(None)
):
    """Get a specific SoldeConge by ID."""
    query = select(SoldeConge).where(SoldeConge.id == solde_id)

    expand_fields = parse_expand_param(expand)
    if expand_fields:
        query = apply_expansion(query, SoldeConge, expand_fields)

    result = await db.execute(query)
    solde = result.scalar_one_or_none()

    if not solde:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Solde de congé non trouvé"
        )

    return SoldeCongeResponse.model_validate(solde)


@router.put("/soldes/{solde_id}", response_model=SoldeCongeResponse)
async def update_solde_conge(
    solde_id: int,
    solde_data: SoldeCongeUpdate,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission("conge", "manage_soldes"))
):
    """Update a SoldeConge with automatic restant recalculation."""
    solde = await SoldeCongeService.update_solde(
        solde_id=solde_id,
        solde_data=solde_data,
        db=db
    )

    await AuditService.log_model_change(
        db=db,
        user=current_user,
        instance=solde,
        action="UPDATE",
        request=request
    )

    return SoldeCongeResponse.model_validate(solde)


@router.delete("/soldes/{solde_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_solde_conge(
    solde_id: int,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission("conge", "manage_soldes"))
):
    """Delete a SoldeConge."""
    query = select(SoldeConge).where(SoldeConge.id == solde_id)
    result = await db.execute(query)
    solde = result.scalar_one_or_none()

    if not solde:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Solde de congé non trouvé"
        )

    old_values = AuditService._extract_model_values(solde)
    await db.delete(solde)
    await db.commit()

    await AuditService.log_action(
        db=db,
        user=current_user,
        action="DELETE",
        resource_type="cg_solde_conge",
        resource_id=str(solde_id),
        old_values=old_values,
        request=request
    )

    return None


@router.post("/soldes/bulk-create")
async def bulk_create_soldes(
    bulk_data: BulkCreateSoldeRequest,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission("conge", "manage_soldes"))
):
    """
    Create leave balances for all employees for a given year and leave type.

    Bulk create leave balances (SoldeConge) for all active employees in the system
    for a specific year and leave type. This is useful for initializing annual leave
    balances at the start of a new year. Skips employees who already have a balance
    for the specified year and leave type.

    **Request Body:**
    ```json
    {
        "annee": 2024,
        "type_conge_id": 1,
        "alloue": 30.0
    }
    ```

    **Example Response:**
    ```json
    {
        "message": "45 soldes créés avec succès",
        "count": 45,
        "soldes": [
            {
                "id": 1,
                "employe_id": 5,
                "type_conge_id": 1,
                "annee": 2024,
                "alloue": 30.0,
                "utilise": 0.0,
                "restant": 30.0
            },
            {
                "id": 2,
                "employe_id": 6,
                "type_conge_id": 1,
                "annee": 2024,
                "alloue": 30.0,
                "utilise": 0.0,
                "restant": 30.0
            }
        ]
    }
    ```

    **Workflow:**
    1. Retrieves all active employees from the system
    2. For each employee, checks if balance already exists
    3. Creates new balance with specified allocated days
    4. Sets used days to 0 and calculates remaining days
    5. Returns list of created balances

    **Use Cases:**
    - Initialize annual leave balances at year start
    - Add new leave type balances for all employees
    - Reset balances for a new fiscal year

    **Error Responses:**
    - **401**: Unauthorized - Invalid or missing authentication
    - **403**: Forbidden - Insufficient permissions (requires 'manage_soldes')
    - **404**: Not Found - Leave type not found
    - **422**: Validation Error - Invalid request body
    """
    created_soldes = await SoldeCongeService.bulk_create_soldes(
        annee=bulk_data.annee,
        type_conge_id=bulk_data.type_conge_id,
        alloue=bulk_data.alloue,
        db=db
    )

    await AuditService.log_action(
        db=db,
        user=current_user,
        action="BULK_CREATE",
        resource_type="cg_solde_conge",
        resource_id=None,
        new_values={
            "annee": bulk_data.annee,
            "type_conge_id": bulk_data.type_conge_id,
            "alloue": bulk_data.alloue,
            "count": len(created_soldes)
        },
        request=request
    )

    return {
        "message": f"{len(created_soldes)} soldes créés avec succès",
        "count": len(created_soldes),
        "soldes": [
            SoldeCongeResponse.model_validate(s) for s in created_soldes
        ]
    }



# ============================================================================
# HistoriqueConge Routes
# ============================================================================

@router.get("/historiques", response_model=PaginatedHistoriqueConge)
async def list_historique_conge(
    db: AsyncSession = Depends(get_db),
    _current_user: User = Depends(require_permission("conge", "view")),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    no_pagination: bool = Query(False),
    ordering: Optional[str] = Query(None),
    expand: Optional[str] = Query(None),
    demande_conge_id: Optional[int] = Query(None),
    valideur_id: Optional[int] = Query(None),
    action: Optional[str] = Query(None),
    niveau_validation: Optional[int] = Query(None)
):
    """List all HistoriqueConge with pagination, filters, and expand support."""
    query = select(HistoriqueConge)

    # Apply filters
    if demande_conge_id:
        query = query.where(HistoriqueConge.demande_conge_id == demande_conge_id)
    if valideur_id:
        query = query.where(HistoriqueConge.valideur_id == valideur_id)
    if action:
        query = query.where(HistoriqueConge.action == action)
    if niveau_validation is not None:
        query = query.where(HistoriqueConge.niveau_validation == niveau_validation)

    # Apply ordering
    if ordering:
        query = apply_ordering(query, HistoriqueConge, ordering)
    else:
        query = query.order_by(HistoriqueConge.date_action.desc())

    # Apply expansion
    expand_fields = parse_expand_param(expand)
    if expand_fields:
        query = apply_expansion(query, HistoriqueConge, expand_fields)

    # Count total
    count_query = select(func.count()).select_from(HistoriqueConge)
    if demande_conge_id:
        count_query = count_query.where(
            HistoriqueConge.demande_conge_id == demande_conge_id
        )
    if valideur_id:
        count_query = count_query.where(HistoriqueConge.valideur_id == valideur_id)
    if action:
        count_query = count_query.where(HistoriqueConge.action == action)
    if niveau_validation is not None:
        count_query = count_query.where(
            HistoriqueConge.niveau_validation == niveau_validation
        )

    result = await db.execute(count_query)
    total = result.scalar()

    # Apply pagination
    if not no_pagination:
        query = query.offset(skip).limit(limit)
    else:
        skip = 0
        limit = total

    result = await db.execute(query)
    historiques = result.scalars().all()

    return PaginatedHistoriqueConge(
        items=[HistoriqueCongeResponse.model_validate(h) for h in historiques],
        total=total,
        skip=skip,
        limit=limit
    )


@router.get("/historiques/{historique_id}", response_model=HistoriqueCongeResponse)
async def get_historique_conge(
    historique_id: int,
    db: AsyncSession = Depends(get_db),
    _current_user: User = Depends(require_permission("conge", "view")),
    expand: Optional[str] = Query(None)
):
    """Get a specific HistoriqueConge by ID."""
    query = select(HistoriqueConge).where(HistoriqueConge.id == historique_id)

    expand_fields = parse_expand_param(expand)
    if expand_fields:
        query = apply_expansion(query, HistoriqueConge, expand_fields)

    result = await db.execute(query)
    historique = result.scalar_one_or_none()

    if not historique:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Historique de congé non trouvé"
        )

    return HistoriqueCongeResponse.model_validate(historique)


# ============================================================================
# Statistics Routes
# ============================================================================

@router.get("/stats", response_model=CongeStatsResponse)
async def get_global_stats(
    db: AsyncSession = Depends(get_db),
    _current_user: User = Depends(require_permission("conge", "view")),
    date_debut: Optional[str] = Query(None),
    date_fin: Optional[str] = Query(None),
    type_conge_id: Optional[int] = Query(None),
    service_id: Optional[int] = Query(None)
):
    """Get global leave statistics with optional filters."""
    query = select(DemandeConge)

    # Apply filters
    if date_debut:
        query = query.where(DemandeConge.date_debut >= date_debut)
    if date_fin:
        query = query.where(DemandeConge.date_fin <= date_fin)
    if type_conge_id:
        query = query.where(DemandeConge.type_conge_id == type_conge_id)

    # Get all demandes matching filters
    result = await db.execute(query)
    demandes = result.scalars().all()

    # Calculate statistics
    total_demandes = len(demandes)

    # Group by status
    demandes_par_statut = {}
    for demande in demandes:
        statut = demande.statut
        demandes_par_statut[statut] = demandes_par_statut.get(statut, 0) + 1

    # Calculate average days per employee
    employe_jours = {}
    for demande in demandes:
        if demande.statut == "APPROVED":
            employe_id = demande.employe_id
            employe_jours[employe_id] = (
                employe_jours.get(employe_id, 0) + demande.nb_jours_ouvrables
            )

    jours_moyens_par_employe = (
        sum(employe_jours.values()) / len(employe_jours)
        if employe_jours else 0.0
    )

    # Calculate utilization rate
    # Get total allocated days for all employees
    solde_query = select(SoldeConge)
    if type_conge_id:
        solde_query = solde_query.where(SoldeConge.type_conge_id == type_conge_id)

    result = await db.execute(solde_query)
    soldes = result.scalars().all()

    total_alloue = sum(s.alloue for s in soldes)
    total_utilise = sum(s.utilise for s in soldes)

    taux_utilisation = (
        (total_utilise / total_alloue * 100) if total_alloue > 0 else 0.0
    )

    return CongeStatsResponse(
        total_demandes=total_demandes,
        demandes_par_statut=demandes_par_statut,
        jours_moyens_par_employe=round(jours_moyens_par_employe, 2),
        taux_utilisation=round(taux_utilisation, 2)
    )


@router.get("/stats/employe/{employe_id}", response_model=EmployeStatsResponse)
async def get_employe_stats(
    employe_id: int,
    db: AsyncSession = Depends(get_db),
    _current_user: User = Depends(require_permission("conge", "view")),
    annee: Optional[int] = Query(None),
    type_conge_id: Optional[int] = Query(None)
):
    """Get leave statistics for a specific employee."""
    # Get all demandes for this employee
    query = select(DemandeConge).where(DemandeConge.employe_id == employe_id)

    if type_conge_id:
        query = query.where(DemandeConge.type_conge_id == type_conge_id)

    result = await db.execute(query)
    demandes = result.scalars().all()

    # Calculate statistics
    total_demandes = len(demandes)
    demandes_approuvees = sum(1 for d in demandes if d.statut == "APPROVED")
    demandes_en_attente = sum(
        1 for d in demandes if d.statut in ["PENDING", "IN_PROGRESS"]
    )
    demandes_rejetees = sum(1 for d in demandes if d.statut == "REJECTED")

    # Get solde information
    solde_query = select(SoldeConge).where(SoldeConge.employe_id == employe_id)

    if annee:
        solde_query = solde_query.where(SoldeConge.annee == annee)
    if type_conge_id:
        solde_query = solde_query.where(SoldeConge.type_conge_id == type_conge_id)

    result = await db.execute(solde_query)
    soldes = result.scalars().all()

    jours_utilises = sum(s.utilise for s in soldes)
    jours_restants = sum(s.restant for s in soldes)
    total_alloue = sum(s.alloue for s in soldes)

    taux_utilisation = (
        (jours_utilises / total_alloue * 100) if total_alloue > 0 else 0.0
    )

    return EmployeStatsResponse(
        employe_id=employe_id,
        total_demandes=total_demandes,
        demandes_approuvees=demandes_approuvees,
        demandes_en_attente=demandes_en_attente,
        demandes_rejetees=demandes_rejetees,
        jours_utilises=round(jours_utilises, 2),
        jours_restants=round(jours_restants, 2),
        taux_utilisation=round(taux_utilisation, 2)
    )


@router.get("/stats/service/{service_id}", response_model=ServiceStatsResponse)
async def get_service_stats(
    service_id: int,
    db: AsyncSession = Depends(get_db),
    _current_user: User = Depends(require_permission("conge", "view")),
    annee: Optional[int] = Query(None),
    type_conge_id: Optional[int] = Query(None)
):
    """Get leave statistics for a specific service."""
    from app.user_app.models import Employe

    # Get all employees in this service
    employe_query = select(Employe).where(Employe.poste_id == service_id)
    result = await db.execute(employe_query)
    employes = result.scalars().all()

    if not employes:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Service non trouvé ou aucun employé dans ce service"
        )

    employe_ids = [e.id for e in employes]
    total_employes = len(employe_ids)

    # Get all demandes for employees in this service
    query = select(DemandeConge).where(DemandeConge.employe_id.in_(employe_ids))

    if type_conge_id:
        query = query.where(DemandeConge.type_conge_id == type_conge_id)

    result = await db.execute(query)
    demandes = result.scalars().all()

    # Calculate statistics
    total_demandes = len(demandes)
    demandes_approuvees = sum(1 for d in demandes if d.statut == "APPROVED")
    demandes_en_attente = sum(
        1 for d in demandes if d.statut in ["PENDING", "IN_PROGRESS"]
    )

    # Get solde information for all employees
    solde_query = select(SoldeConge).where(SoldeConge.employe_id.in_(employe_ids))

    if annee:
        solde_query = solde_query.where(SoldeConge.annee == annee)
    if type_conge_id:
        solde_query = solde_query.where(SoldeConge.type_conge_id == type_conge_id)

    result = await db.execute(solde_query)
    soldes = result.scalars().all()

    total_utilise = sum(s.utilise for s in soldes)
    total_alloue = sum(s.alloue for s in soldes)

    jours_moyens_par_employe = (
        total_utilise / total_employes if total_employes > 0 else 0.0
    )
    taux_utilisation_moyen = (
        (total_utilise / total_alloue * 100) if total_alloue > 0 else 0.0
    )

    return ServiceStatsResponse(
        service_id=service_id,
        total_employes=total_employes,
        total_demandes=total_demandes,
        demandes_approuvees=demandes_approuvees,
        demandes_en_attente=demandes_en_attente,
        jours_moyens_par_employe=round(jours_moyens_par_employe, 2),
        taux_utilisation_moyen=round(taux_utilisation_moyen, 2)
    )

